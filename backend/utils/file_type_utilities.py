"""
File Type Detection and Handling Utilities

This module provides utilities for detecting and handling different file types,
including MIME type detection, file format validation, and content extraction.

Dependencies:
    - python-magic: Optional library for more accurate MIME type detection. 
      If not available, fallbacks to mimetypes module and basic content inspection.
    - pandas: Required for data loading and conversion between formats.
    - openpyxl: Required for Excel file handling.
    - PIL: Required for image file handling.
    - yaml: Required for YAML file handling.
"""

from __future__ import annotations

import os
import io
import re
import json
import csv
import mimetypes
import logging
import uuid
import tempfile
import zipfile
import tarfile
from functools import lru_cache
from typing import Dict, Any, Optional, List, Union, BinaryIO, Tuple
from datetime import datetime

import pandas as pd
import openpyxl
import xml.etree.ElementTree as ET
from PIL import Image, UnidentifiedImageError
import yaml
try:
    import magic
    MAGIC_AVAILABLE = True
except ImportError:
    MAGIC_AVAILABLE = False

logger = logging.getLogger(__name__)

# Initialize mime types
mimetypes.init()

# Constants for MIME type groups
TABULAR_MIME_TYPES = [
    'text/csv', 
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/json'
]

IMAGE_MIME_TYPES = [
    'image/png',
    'image/jpeg',
    'image/gif',
    'image/bmp',
    'image/tiff'
]

ARCHIVE_MIME_TYPES = [
    'application/zip',
    'application/x-zip-compressed',
    'application/gzip',
    'application/x-gzip'
]

# File signature patterns
FILE_SIGNATURES = {
    'image/png': b'\x89PNG\r\n\x1a\n',
    'image/jpeg': b'\xff\xd8',
    'image/gif': [b'GIF87a', b'GIF89a'],
    'application/pdf': b'%PDF',
    'application/zip': b'PK\x03\x04',
    'application/gzip': b'\x1f\x8b',
    'application/xml': b'<?xml'
}


class FileTypeError(Exception):
    """
    Exception raised for file type errors
    
    This is the base exception class for all file type related errors in this module.
    It includes errors related to file detection, conversion, validation, and data extraction.
    """
    pass


@lru_cache(maxsize=128)
def detect_mime_type(file_path: str = None, file_data: bytes = None, file_name: str = None) -> str:
    """
    Detect the MIME type of a file
    
    Args:
        file_path: Path to the file (optional)
        file_data: File data as bytes (optional)
        file_name: File name for extension-based detection (optional)
        
    Returns:
        str: MIME type of the file
    
    Note:
        At least one of file_path, file_data, or file_name must be provided.
    """
    if not file_path and not file_data and not file_name:
        raise ValueError("At least one of file_path, file_data, or file_name must be provided")
    
    # Use python-magic if available (more accurate)
    if MAGIC_AVAILABLE and (file_path or file_data):
        try:
            if file_path:
                return magic.from_file(file_path, mime=True)
            else:
                return magic.from_buffer(file_data, mime=True)
        except Exception as e:
            logger.warning(f"Error using python-magic: {str(e)}")
            # Fall back to other methods
    
    # Use mimetypes module with file name/extension
    if file_path or file_name:
        name = file_path if file_path else file_name
        mime_type, _ = mimetypes.guess_type(name)
        if mime_type:
            return mime_type
    
    # Basic content-based detection if file_data is provided
    if file_data:
        # Check for common file signatures using our constants
        for mime_type, signature in FILE_SIGNATURES.items():
            if isinstance(signature, list):
                # Handle multiple possible signatures (like GIF)
                for sig in signature:
                    if file_data[:len(sig)] == sig:
                        return mime_type
            elif file_data[:len(signature)] == signature:
                return mime_type
        
        # Check if it's text
        try:
            text_sample = file_data[:1024].decode('utf-8')
            
            # Check for JSON
            try:
                json.loads(text_sample)
                return 'application/json'
            except:
                pass
            
            # Check for CSV (look for consistent delimiters)
            if ',' in text_sample and '\n' in text_sample:
                lines = text_sample.split('\n')
                if len(lines) > 1:
                    comma_counts = [line.count(',') for line in lines[:5] if line.strip()]
                    if len(set(comma_counts)) == 1 and comma_counts[0] > 0:
                        return 'text/csv'
            
            # Check for YAML
            if text_sample.lstrip().startswith(('---', 'apiVersion:', 'kind:')):
                try:
                    yaml.safe_load(text_sample)
                    return 'application/yaml'
                except:
                    pass
            
            # Generic text
            return 'text/plain'
        except:
            pass
    
    # Default fallback
    return 'application/octet-stream'


@lru_cache(maxsize=128)
def get_file_extension(mime_type: str) -> str:
    """
    Get the appropriate file extension for a MIME type
    
    Args:
        mime_type: MIME type
        
    Returns:
        str: File extension including the dot (e.g., '.txt')
    """
    extension = mimetypes.guess_extension(mime_type)
    
    # Handle special cases where the default extension isn't ideal
    if mime_type == 'application/json' and extension != '.json':
        extension = '.json'
    elif mime_type == 'text/csv' and extension != '.csv':
        extension = '.csv'
    elif mime_type == 'application/yaml' and extension != '.yaml':
        extension = '.yaml'
    
    return extension or ''


def convert_to_standard_format(file_data: bytes, source_mime_type: str, target_format: str) -> Tuple[bytes, str]:
    """
    Convert file data from one format to another
    
    Args:
        file_data: File data as bytes
        source_mime_type: Source MIME type
        target_format: Target format name ('csv', 'json', 'excel', etc.)
        
    Returns:
        Tuple[bytes, str]: Converted data and the new MIME type
    """
    # Map target formats to MIME types
    format_to_mime = {
        'csv': 'text/csv',
        'json': 'application/json',
        'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'xml': 'application/xml',
        'yaml': 'application/yaml',
        'text': 'text/plain'
    }
    
    target_mime = format_to_mime.get(target_format.lower())
    if not target_mime:
        raise ValueError(f"Unsupported target format: {target_format}")
    
    # Return as-is if already in the target format
    if source_mime_type == target_mime:
        return file_data, source_mime_type
    
    # Create a temporary file for processing
    temp_file = None
    try:
        # Load data into memory using pandas if it's tabular
        if source_mime_type in TABULAR_MIME_TYPES:
            # Load source data into pandas DataFrame
            df = load_dataframe(file_data, source_mime_type)
            
            # Convert to target format
            buffer = io.BytesIO()
            
            if target_mime == 'text/csv':
                df.to_csv(buffer, index=False)
            elif target_mime == 'application/json':
                df.to_json(buffer, orient='records')
            elif target_mime == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                df.to_excel(buffer, index=False, engine='openpyxl')
            elif target_mime == 'application/xml':
                xml_str = df.to_xml()
                buffer.write(xml_str.encode('utf-8'))
            elif target_mime == 'application/yaml':
                # Convert to list of dicts first
                records = df.to_dict(orient='records')
                yaml_str = yaml.dump(records, default_flow_style=False)
                buffer.write(yaml_str.encode('utf-8'))
            elif target_mime == 'text/plain':
                buffer.write(df.to_string(index=False).encode('utf-8'))
            
            buffer.seek(0)
            converted_data = buffer.read()
            
            return converted_data, target_mime
            
        elif source_mime_type == 'application/xml' and target_mime in ['application/json', 'text/csv', 'application/yaml']:
            # Parse XML
            root = ET.fromstring(file_data.decode('utf-8'))
            
            # Convert to dict
            def xml_to_dict(element):
                result = {}
                for child in element:
                    child_data = xml_to_dict(child)
                    if child.tag in result:
                        if isinstance(result[child.tag], list):
                            result[child.tag].append(child_data if child_data else child.text)
                        else:
                            result[child.tag] = [result[child.tag], child_data if child_data else child.text]
                    else:
                        result[child.tag] = child_data if child_data else child.text
                return result if element else None
            
            data_dict = xml_to_dict(root)
            
            # Convert to target format
            if target_mime == 'application/json':
                result = json.dumps(data_dict).encode('utf-8')
            elif target_mime == 'text/csv':
                # Flatten the structure (simple case, will lose hierarchy)
                df = pd.json_normalize(data_dict)
                buffer = io.StringIO()
                df.to_csv(buffer, index=False)
                result = buffer.getvalue().encode('utf-8')
            elif target_mime == 'application/yaml':
                result = yaml.dump(data_dict, default_flow_style=False).encode('utf-8')
            
            return result, target_mime
            
        elif source_mime_type == 'application/yaml' and target_mime in ['application/json', 'text/csv', 'application/xml']:
            # Parse YAML
            data_dict = yaml.safe_load(file_data)
            
            # Convert to target format
            if target_mime == 'application/json':
                result = json.dumps(data_dict).encode('utf-8')
            elif target_mime == 'text/csv':
                # Convert to DataFrame if possible
                try:
                    df = pd.DataFrame(data_dict)
                    buffer = io.StringIO()
                    df.to_csv(buffer, index=False)
                    result = buffer.getvalue().encode('utf-8')
                except:
                    # Try to normalize JSON-like structure
                    df = pd.json_normalize(data_dict)
                    buffer = io.StringIO()
                    df.to_csv(buffer, index=False)
                    result = buffer.getvalue().encode('utf-8')
            elif target_mime == 'application/xml':
                # Simple conversion using ElementTree
                def dict_to_xml(d, root_name='root'):
                    root = ET.Element(root_name)
                    
                    def _dict_to_xml(d, parent):
                        for key, value in d.items():
                            if isinstance(value, dict):
                                child = ET.SubElement(parent, key)
                                _dict_to_xml(value, child)
                            elif isinstance(value, list):
                                for item in value:
                                    child = ET.SubElement(parent, key)
                                    if isinstance(item, dict):
                                        _dict_to_xml(item, child)
                                    else:
                                        child.text = str(item)
                            else:
                                child = ET.SubElement(parent, key)
                                child.text = str(value)
                                
                    _dict_to_xml(d, root)
                    return root
                
                root = dict_to_xml(data_dict)
                result = ET.tostring(root, encoding='utf-8')
            
            return result, target_mime
            
        else:
            raise ValueError(f"Conversion from {source_mime_type} to {target_mime} is not supported")
    
    except Exception as e:
        logger.error(f"Error during file format conversion: {str(e)}")
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception as cleanup_error:
                logger.warning(f"Failed to remove temporary file {temp_file}: {cleanup_error}")
        raise FileTypeError(f"Failed to convert from {source_mime_type} to {target_format}: {str(e)}")
            
    finally:
        # Clean up temporary file if created
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception as cleanup_error:
                logger.warning(f"Failed to remove temporary file {temp_file}: {cleanup_error}")


def extract_metadata(file_data: bytes, mime_type: str) -> Dict[str, Any]:
    """
    Extract metadata from a file
    
    Args:
        file_data: File data as bytes
        mime_type: MIME type
        
    Returns:
        Dict[str, Any]: Metadata dictionary
    """
    metadata = {}
    
    # Create a temporary file for processing
    temp_file = None
    try:
        # Extract metadata based on file type
        if mime_type.startswith('image/'):
            # Image metadata
            try:
                img = Image.open(io.BytesIO(file_data))
                metadata['format'] = img.format
                metadata['mode'] = img.mode
                metadata['width'] = img.width
                metadata['height'] = img.height
                
                # Extract EXIF data if available
                if hasattr(img, '_getexif') and img._getexif():
                    exif_data = img._getexif()
                    if exif_data:
                        metadata['exif'] = {}
                        for tag_id, value in exif_data.items():
                            # Skip binary data
                            if isinstance(value, bytes):
                                continue
                            metadata['exif'][str(tag_id)] = str(value)
            except Exception as e:
                logger.warning(f"Error extracting image metadata: {str(e)}")
        
        elif mime_type == 'application/pdf':
            # PDF metadata - would need PyPDF2 or similar library
            # Here, just extract basic info
            if file_data[:4] == b'%PDF':
                version_match = re.search(rb'%PDF-(\d+\.\d+)', file_data[:20])
                if version_match:
                    metadata['pdf_version'] = version_match.group(1).decode('utf-8')
        
        elif mime_type in ['application/vnd.ms-excel', 
                           'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            # Excel metadata
            try:
                temp_file = tempfile.mktemp(suffix='.xlsx')
                with open(temp_file, 'wb') as f:
                    f.write(file_data)
                
                wb = openpyxl.load_workbook(temp_file, read_only=True)
                metadata['sheet_names'] = wb.sheetnames
                metadata['sheet_count'] = len(wb.sheetnames)
                
                # Extract document properties
                props = wb.properties
                if props:
                    if props.title:
                        metadata['title'] = props.title
                    if props.creator:
                        metadata['creator'] = props.creator
                    if props.created:
                        metadata['created'] = props.created.isoformat()
                    if props.modified:
                        metadata['modified'] = props.modified.isoformat()
            except Exception as e:
                logger.warning(f"Error extracting Excel metadata: {str(e)}")
        
        elif mime_type == 'text/csv':
            # CSV metadata
            try:
                # Try to determine delimiter and count rows/columns
                text_data = file_data.decode('utf-8')
                dialect = csv.Sniffer().sniff(text_data[:4096])
                metadata['delimiter'] = dialect.delimiter
                
                # Count rows and columns
                lines = text_data.splitlines()
                metadata['row_count'] = len(lines)
                if lines:
                    metadata['column_count'] = len(next(csv.reader([lines[0]], dialect)))
            except Exception as e:
                logger.warning(f"Error extracting CSV metadata: {str(e)}")
        
        elif mime_type == 'application/json':
            # JSON metadata
            try:
                json_data = json.loads(file_data)
                if isinstance(json_data, dict):
                    metadata['keys'] = list(json_data.keys())
                    metadata['top_level_type'] = 'object'
                elif isinstance(json_data, list):
                    metadata['length'] = len(json_data)
                    metadata['top_level_type'] = 'array'
                    if json_data and isinstance(json_data[0], dict):
                        metadata['sample_keys'] = list(json_data[0].keys())
            except Exception as e:
                logger.warning(f"Error extracting JSON metadata: {str(e)}")
        
        elif mime_type == 'application/xml':
            # XML metadata
            try:
                root = ET.fromstring(file_data.decode('utf-8'))
                metadata['root_tag'] = root.tag
                
                # Count elements
                metadata['element_count'] = len(root.findall('.//*'))
                
                # Get namespaces
                namespaces = dict([node for _, node in ET.iterparse(io.BytesIO(file_data), events=['start-ns'])])
                if namespaces:
                    metadata['namespaces'] = namespaces
            except Exception as e:
                logger.warning(f"Error extracting XML metadata: {str(e)}")
        
        elif mime_type in ['application/zip', 'application/x-zip-compressed']:
            # ZIP metadata
            try:
                with zipfile.ZipFile(io.BytesIO(file_data)) as z:
                    metadata['file_count'] = len(z.namelist())
                    metadata['files'] = z.namelist()
                    
                    # Get information about the first few files
                    metadata['file_info'] = []
                    for i, info in enumerate(z.infolist()):
                        if i >= 5:  # Limit to first 5 files
                            break
                        metadata['file_info'].append({
                            'filename': info.filename,
                            'file_size': info.file_size,
                            'compress_size': info.compress_size,
                            'date_time': datetime(*info.date_time).isoformat()
                        })
            except Exception as e:
                logger.warning(f"Error extracting ZIP metadata: {str(e)}")
        
        elif mime_type in ['application/gzip', 'application/x-gzip']:
            # GZIP metadata - basic info only
            if file_data[:2] == b'\x1f\x8b':
                metadata['gzip_format'] = True
                
                # Try to extract tar archive info if it's a tar.gz
                try:
                    with tarfile.open(fileobj=io.BytesIO(file_data), mode='r:gz') as tar:
                        members = tar.getmembers()
                        metadata['file_count'] = len(members)
                        metadata['files'] = [m.name for m in members]
                except Exception as e:
                    logger.warning(f"Error extracting tar.gz metadata: {str(e)}")
    
    except Exception as e:
        logger.error(f"Error extracting metadata: {str(e)}")
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception as cleanup_error:
                logger.warning(f"Failed to remove temporary file {temp_file}: {cleanup_error}")
        # Return partial metadata if we have any
        if not metadata:
            metadata = {'error': str(e)}
    
    finally:
        # Clean up temporary file if created
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception as cleanup_error:
                logger.warning(f"Failed to remove temporary file {temp_file}: {cleanup_error}")
    
    return metadata


def validate_file_format(file_data: bytes, mime_type: str) -> bool:
    """
    Validate if file data matches the expected MIME type format
    
    Args:
        file_data: File data as bytes
        mime_type: Expected MIME type
        
    Returns:
        bool: True if valid, False otherwise
    """
    try:
        # Basic validation based on file signatures and parsing
        if mime_type in FILE_SIGNATURES:
            signature = FILE_SIGNATURES[mime_type]
            if isinstance(signature, list):
                # Check multiple signatures (like GIF)
                valid = False
                for sig in signature:
                    if file_data[:len(sig)] == sig:
                        valid = True
                        break
                if not valid:
                    return False
            elif not file_data[:len(signature)].startswith(signature):
                return False
        
        # More sophisticated validation by attempting to parse/process
        if mime_type in IMAGE_MIME_TYPES:
            try:
                Image.open(io.BytesIO(file_data)).verify()
                return True
            except Exception as e:
                logger.debug(f"Image validation failed: {str(e)}")
                return False
        
        elif mime_type == 'application/json':
            try:
                json.loads(file_data)
                return True
            except Exception as e:
                logger.debug(f"JSON validation failed: {str(e)}")
                return False
        
        elif mime_type == 'text/csv':
            try:
                # Try to parse as CSV
                text_data = file_data.decode('utf-8')
                dialect = csv.Sniffer().sniff(text_data[:4096])
                csv.reader(io.StringIO(text_data), dialect)
                return True
            except Exception as e:
                logger.debug(f"CSV validation failed: {str(e)}")
                return False
        
        elif mime_type == 'application/xml':
            try:
                ET.fromstring(file_data.decode('utf-8'))
                return True
            except Exception as e:
                logger.debug(f"XML validation failed: {str(e)}")
                return False
        
        elif mime_type == 'application/yaml':
            try:
                yaml.safe_load(file_data)
                return True
            except Exception as e:
                logger.debug(f"YAML validation failed: {str(e)}")
                return False
        
        elif 'excel' in mime_type.lower() or (mime_type in TABULAR_MIME_TYPES and 'sheet' in mime_type.lower()):
            temp_file = None
            try:
                # Write to temp file for validation
                temp_file = tempfile.mktemp(suffix='.xlsx')
                with open(temp_file, 'wb') as f:
                    f.write(file_data)
                
                openpyxl.load_workbook(temp_file, read_only=True)
                return True
            except Exception as e:
                logger.warning(f"Error validating Excel file: {str(e)}")
                return False
            finally:
                # Clean up temp file
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except Exception as cleanup_error:
                        logger.warning(f"Failed to remove temporary file {temp_file}: {cleanup_error}")
        
        elif mime_type in ARCHIVE_MIME_TYPES and ('zip' in mime_type.lower()):
            try:
                with zipfile.ZipFile(io.BytesIO(file_data)) as zip_file:
                    test_result = zip_file.testzip()
                    if test_result is not None:
                        logger.warning(f"ZIP validation found corrupted file: {test_result}")
                        return False
                    return True
            except Exception as e:
                logger.debug(f"ZIP validation failed: {str(e)}")
                return False
        
        # For other types, assume valid if we've reached this point
        return True
    
    except Exception as e:
        logger.warning(f"Error validating file format: {str(e)}")
        return False


@lru_cache(maxsize=64)
def load_dataframe(file_data: bytes, mime_type: str) -> pd.DataFrame:
    """
    Load file data into a pandas DataFrame
    
    Args:
        file_data: File data as bytes
        mime_type: MIME type
        
    Returns:
        pd.DataFrame: Loaded data as DataFrame
    """
    try:
        if mime_type == 'text/csv':
            return pd.read_csv(io.BytesIO(file_data))
        
        elif 'excel' in mime_type.lower() or (mime_type in TABULAR_MIME_TYPES and 'sheet' in mime_type.lower()):
            return pd.read_excel(io.BytesIO(file_data))
        
        elif mime_type == 'application/json':
            return pd.read_json(io.BytesIO(file_data))
        
        elif mime_type == 'application/xml':
            return pd.read_xml(io.BytesIO(file_data))
        
        elif mime_type == 'text/tab-separated-values':
            return pd.read_csv(io.BytesIO(file_data), sep='\t')
        
        elif mime_type == 'application/yaml':
            data = yaml.safe_load(file_data)
            if isinstance(data, list):
                return pd.DataFrame(data)
            else:
                # Convert dict to single-row DataFrame
                return pd.DataFrame([data])
        
        else:
            raise ValueError(f"Unsupported file type for DataFrame loading: {mime_type}")
    
    except Exception as e:
        logger.error(f"Error loading DataFrame: {str(e)}")
        raise FileTypeError(f"Failed to load data as DataFrame: {str(e)}")


def generate_unique_filename(original_filename: str = None, mime_type: str = None) -> str:
    """
    Generate a unique filename
    
    Args:
        original_filename: Original filename (optional)
        mime_type: MIME type (optional)
        
    Returns:
        str: Unique filename
    """
    # Generate unique ID
    unique_id = uuid.uuid4().hex
    
    # Try to keep original extension if filename provided
    extension = ''
    if original_filename:
        _, extension = os.path.splitext(original_filename)
    
    # Use MIME type to determine extension if not derived from filename
    if not extension and mime_type:
        extension = get_file_extension(mime_type)
    
    # Default to .bin if no extension determined
    if not extension:
        extension = '.bin'
    
    return f"{unique_id}{extension}"